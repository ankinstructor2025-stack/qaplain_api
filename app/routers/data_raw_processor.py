
import csv
import io
import json
import math
import re
import uuid
import xml.etree.ElementTree as ET
from datetime import date, datetime
from email import policy
from email.parser import BytesParser
from pathlib import PurePosixPath
from typing import Any, Iterable

from fastapi import HTTPException
from google.cloud import firestore
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.firebase import get_firestore_client
from app.routers.data_import_common import (
    get_storage_bucket,
    normalize_extension,
    normalize_text,
    now_iso,
)


SUPPORTED_FILE_TYPE_COLLECTION = "supported_file_types"
DATA_IMPORT_COLLECTION = "data_import"
DATA_SOURCE_COLLECTION = "data_sources"

RAW_DOCUMENT_COLLECTION = "raw_documents"
RAW_RECORD_SUBCOLLECTION = "records"

MAX_TEXT_LENGTH = 180_000
BATCH_WRITE_LIMIT = 400

HEADING_PATTERNS = (
    re.compile(r"^\s*第[0-9０-９一二三四五六七八九十百千]+[章節款項]\s*.+$"),
    re.compile(r"^\s*[0-9０-９]+(?:[.\-．－][0-9０-９]+)*[.．、)]?\s+.+$"),
    re.compile(r"^\s*[（(]?[0-9０-９一二三四五六七八九十]+[）)]\s*.+$"),
    re.compile(r"^\s*[■◆●○◎◇□]\s*.+$"),
)


def _safe_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(value, (str, int, bool)):
        return value

    if isinstance(value, dict):
        return {
            str(key): _safe_value(item)
            for key, item in value.items()
        }

    if isinstance(value, (list, tuple, set)):
        return [
            _safe_value(item)
            for item in value
        ]

    return str(value)


def _split_long_text(
    text: str,
    max_length: int = MAX_TEXT_LENGTH,
) -> list[str]:
    normalized = normalize_text(text)
    if not normalized:
        return [""]

    if len(normalized) <= max_length:
        return [normalized]

    chunks = []
    current = []

    for paragraph in re.split(r"\n{2,}", normalized):
        paragraph = paragraph.strip()

        if not paragraph:
            continue

        candidate = "\n\n".join(
            current + [paragraph]
        )

        if (
            current
            and len(candidate) > max_length
        ):
            chunks.append(
                "\n\n".join(current)
            )
            current = [paragraph]
        else:
            current.append(paragraph)

        while (
            current
            and len(current[0]) > max_length
        ):
            long_text = current.pop(0)
            chunks.append(
                long_text[:max_length]
            )
            remaining = long_text[max_length:]

            if remaining:
                current.insert(0, remaining)

    if current:
        chunks.append(
            "\n\n".join(current)
        )

    return chunks or [normalized[:max_length]]


def _looks_like_heading(line: str) -> bool:
    value = normalize_text(line)

    if not value:
        return False

    if len(value) > 120:
        return False

    if value.endswith(("。", "、", "，", ",")):
        return False

    return any(
        pattern.match(value)
        for pattern in HEADING_PATTERNS
    )


def _split_lines_by_heading(
    lines: list[tuple[int | None, str]],
) -> list[dict]:
    sections = []
    current_title = ""
    current_lines = []
    start_page = None
    end_page = None

    def flush() -> None:
        nonlocal current_title
        nonlocal current_lines
        nonlocal start_page
        nonlocal end_page

        content = "\n".join(
            line
            for _, line in current_lines
        ).strip()

        if not content and not current_title:
            return

        sections.append({
            "title": current_title,
            "content": content,
            "start_page": start_page,
            "end_page": end_page,
        })

        current_title = ""
        current_lines = []
        start_page = None
        end_page = None

    for page_number, line in lines:
        clean_line = line.strip()

        if not clean_line:
            if current_lines:
                current_lines.append(
                    (page_number, "")
                )
            continue

        if _looks_like_heading(clean_line):
            flush()
            current_title = clean_line
            start_page = page_number
            end_page = page_number
            continue

        if start_page is None:
            start_page = page_number

        end_page = page_number
        current_lines.append(
            (page_number, clean_line)
        )

    flush()

    return sections


def _record(
    *,
    record_type: str,
    title: str = "",
    content: str = "",
    sequence: int = 0,
    metadata: dict | None = None,
    start_page: int | None = None,
    end_page: int | None = None,
    structured_data: Any = None,
) -> dict:
    return {
        "record_type": record_type,
        "title": normalize_text(title),
        "content": normalize_text(content),
        "sequence": sequence,
        "start_page": start_page,
        "end_page": end_page,
        "structured_data": _safe_value(
            structured_data
        ),
        "metadata": _safe_value(
            metadata or {}
        ),
    }


def _records_from_json(content: bytes) -> list[dict]:
    try:
        data = json.loads(
            content.decode(
                "utf-8-sig"
            )
        )
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                "JSONファイルを解析できません。"
                f" {type(error).__name__}: {error}"
            ),
        )

    if isinstance(data, list):
        values = data
        root_name = "root"
    elif isinstance(data, dict):
        list_candidates = [
            (key, value)
            for key, value in data.items()
            if isinstance(value, list)
        ]

        if len(list_candidates) == 1:
            root_name, values = list_candidates[0]
        else:
            return [
                _record(
                    record_type="json_object",
                    title="root",
                    sequence=1,
                    structured_data=data,
                    content=json.dumps(
                        data,
                        ensure_ascii=False,
                        separators=(",", ":"),
                    ),
                )
            ]
    else:
        return [
            _record(
                record_type="json_value",
                title="root",
                sequence=1,
                structured_data=data,
                content=json.dumps(
                    data,
                    ensure_ascii=False,
                ),
            )
        ]

    records = []

    for index, value in enumerate(
        values,
        start=1,
    ):
        title = ""

        if isinstance(value, dict):
            for key in (
                "title",
                "name",
                "subject",
                "id",
                "record_id",
            ):
                if value.get(key) is not None:
                    title = normalize_text(
                        value.get(key)
                    )
                    if title:
                        break

        records.append(
            _record(
                record_type="json_item",
                title=title or f"{root_name}[{index - 1}]",
                sequence=index,
                structured_data=value,
                content=json.dumps(
                    value,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                metadata={
                    "root_name": root_name,
                    "array_index": index - 1,
                },
            )
        )

    return records


def _xml_element_to_data(
    element: ET.Element,
) -> dict:
    children = list(element)

    data = {
        "tag": element.tag,
        "attributes": dict(
            element.attrib
        ),
        "text": normalize_text(
            element.text
        ),
    }

    if children:
        data["children"] = [
            _xml_element_to_data(child)
            for child in children
        ]

    return data


def _xml_text(
    element: ET.Element,
) -> str:
    values = [
        normalize_text(value)
        for value in element.itertext()
        if normalize_text(value)
    ]
    return "\n".join(values)


def _records_from_xml(content: bytes) -> list[dict]:
    try:
        root = ET.fromstring(content)
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                "XMLファイルを解析できません。"
                f" {type(error).__name__}: {error}"
            ),
        )

    children = list(root)

    if not children:
        children = [root]

    records = []

    for index, element in enumerate(
        children,
        start=1,
    ):
        records.append(
            _record(
                record_type="xml_element",
                title=element.tag,
                content=_xml_text(element),
                sequence=index,
                structured_data=(
                    _xml_element_to_data(
                        element
                    )
                ),
                metadata={
                    "root_tag": root.tag,
                    "element_tag": element.tag,
                },
            )
        )

    return records


def _detect_csv_encoding(
    content: bytes,
) -> str:
    for encoding in (
        "utf-8-sig",
        "utf-8",
        "cp932",
    ):
        try:
            content.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    return "utf-8"


def _records_from_csv(content: bytes) -> list[dict]:
    encoding = _detect_csv_encoding(content)
    text = content.decode(
        encoding,
        errors="replace",
    )

    sample = text[:8192]

    try:
        dialect = csv.Sniffer().sniff(
            sample,
            delimiters=",\t;",
        )
    except Exception:
        dialect = csv.excel

    reader = csv.DictReader(
        io.StringIO(text),
        dialect=dialect,
    )

    headers = [
        normalize_text(header)
        for header in (
            reader.fieldnames or []
        )
    ]

    records = []

    for index, row in enumerate(
        reader,
        start=1,
    ):
        normalized_row = {
            normalize_text(key):
                normalize_text(value)
            for key, value in row.items()
            if key is not None
        }

        records.append(
            _record(
                record_type="csv_row",
                title=f"行 {index + 1}",
                sequence=index,
                structured_data=normalized_row,
                content=json.dumps(
                    normalized_row,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                metadata={
                    "row_number": index + 1,
                    "headers": headers,
                    "encoding": encoding,
                },
            )
        )

    return records


def _records_from_xlsx(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
        )
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                "XLSXファイルを解析できません。"
                f" {type(error).__name__}: {error}"
            ),
        )

    records = []
    sequence = 0

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(
                values_only=True
            )

            try:
                first_row = next(rows)
            except StopIteration:
                continue

            headers = [
                normalize_text(value)
                or f"column_{index}"
                for index, value in enumerate(
                    first_row,
                    start=1,
                )
            ]

            for row_number, values in enumerate(
                rows,
                start=2,
            ):
                if not any(
                    value is not None
                    and normalize_text(value)
                    for value in values
                ):
                    continue

                sequence += 1

                row_data = {
                    headers[index]:
                        _safe_value(value)
                    for index, value in enumerate(
                        values
                    )
                    if index < len(headers)
                }

                records.append(
                    _record(
                        record_type="xlsx_row",
                        title=(
                            f"{worksheet.title}"
                            f" / 行 {row_number}"
                        ),
                        sequence=sequence,
                        structured_data=row_data,
                        content=json.dumps(
                            row_data,
                            ensure_ascii=False,
                            separators=(",", ":"),
                        ),
                        metadata={
                            "sheet_name":
                                worksheet.title,
                            "row_number":
                                row_number,
                            "headers":
                                headers,
                        },
                    )
                )
    finally:
        workbook.close()

    return records


def _records_from_txt(content: bytes) -> list[dict]:
    encoding = _detect_csv_encoding(content)
    text = content.decode(
        encoding,
        errors="replace",
    )

    lines = [
        (None, line)
        for line in text.splitlines()
    ]

    sections = _split_lines_by_heading(
        lines
    )

    heading_count = sum(
        1
        for section in sections
        if section["title"]
    )

    records = []

    if heading_count >= 2:
        for index, section in enumerate(
            sections,
            start=1,
        ):
            chunks = _split_long_text(
                section["content"]
            )

            for chunk_index, chunk in enumerate(
                chunks,
                start=1,
            ):
                title = section["title"]

                if len(chunks) > 1:
                    title = (
                        f"{title} "
                        f"({chunk_index}/{len(chunks)})"
                    )

                records.append(
                    _record(
                        record_type="text_section",
                        title=title,
                        content=chunk,
                        sequence=len(records) + 1,
                        metadata={
                            "encoding": encoding,
                            "split_method":
                                "heading",
                        },
                    )
                )

        return records

    for index, chunk in enumerate(
        _split_long_text(text),
        start=1,
    ):
        records.append(
            _record(
                record_type="text_chunk",
                title=f"本文 {index}",
                content=chunk,
                sequence=index,
                metadata={
                    "encoding": encoding,
                    "split_method":
                        "fixed_length",
                },
            )
        )

    return records


def _records_from_pdf(content: bytes) -> tuple[list[dict], dict]:
    try:
        reader = PdfReader(
            io.BytesIO(content)
        )
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                "PDFファイルを解析できません。"
                f" {type(error).__name__}: {error}"
            ),
        )

    page_lines = []
    page_texts = []

    for page_number, page in enumerate(
        reader.pages,
        start=1,
    ):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""

        page_texts.append(text)

        for line in text.splitlines():
            page_lines.append(
                (page_number, line)
            )

    sections = _split_lines_by_heading(
        page_lines
    )

    heading_count = sum(
        1
        for section in sections
        if section["title"]
    )

    records = []

    if heading_count >= 2:
        for section in sections:
            chunks = _split_long_text(
                section["content"]
            )

            for chunk_index, chunk in enumerate(
                chunks,
                start=1,
            ):
                title = section["title"]

                if len(chunks) > 1:
                    title = (
                        f"{title} "
                        f"({chunk_index}/{len(chunks)})"
                    )

                records.append(
                    _record(
                        record_type="pdf_section",
                        title=title,
                        content=chunk,
                        sequence=len(records) + 1,
                        start_page=(
                            section["start_page"]
                        ),
                        end_page=(
                            section["end_page"]
                        ),
                        metadata={
                            "split_method":
                                "heading",
                        },
                    )
                )
    else:
        for page_number, text in enumerate(
            page_texts,
            start=1,
        ):
            chunks = _split_long_text(text)

            for chunk_index, chunk in enumerate(
                chunks,
                start=1,
            ):
                title = f"{page_number}ページ"

                if len(chunks) > 1:
                    title = (
                        f"{title} "
                        f"({chunk_index}/{len(chunks)})"
                    )

                records.append(
                    _record(
                        record_type="pdf_page",
                        title=title,
                        content=chunk,
                        sequence=len(records) + 1,
                        start_page=page_number,
                        end_page=page_number,
                        metadata={
                            "split_method":
                                "page",
                        },
                    )
                )

    metadata = {
        "page_count": len(reader.pages),
        "pdf_metadata": _safe_value(
            dict(reader.metadata or {})
        ),
        "split_method": (
            "heading"
            if heading_count >= 2
            else "page"
        ),
    }

    return records, metadata


def _extract_email_body(
    message,
) -> tuple[str, str]:
    plain_parts = []
    html_parts = []

    if message.is_multipart():
        for part in message.walk():
            disposition = (
                part.get_content_disposition()
            )

            if disposition == "attachment":
                continue

            content_type = part.get_content_type()

            try:
                value = part.get_content()
            except Exception:
                payload = part.get_payload(
                    decode=True
                ) or b""
                charset = (
                    part.get_content_charset()
                    or "utf-8"
                )
                value = payload.decode(
                    charset,
                    errors="replace",
                )

            if content_type == "text/plain":
                plain_parts.append(str(value))
            elif content_type == "text/html":
                html_parts.append(str(value))
    else:
        try:
            value = message.get_content()
        except Exception:
            value = ""

        if message.get_content_type() == "text/html":
            html_parts.append(str(value))
        else:
            plain_parts.append(str(value))

    return (
        "\n".join(plain_parts).strip(),
        "\n".join(html_parts).strip(),
    )


def _records_from_eml(content: bytes) -> tuple[list[dict], dict]:
    try:
        message = BytesParser(
            policy=policy.default
        ).parsebytes(content)
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                "EMLファイルを解析できません。"
                f" {type(error).__name__}: {error}"
            ),
        )

    plain_body, html_body = (
        _extract_email_body(message)
    )

    body = plain_body or html_body

    attachments = []

    for part in message.iter_attachments():
        payload = part.get_payload(
            decode=True
        ) or b""

        attachments.append({
            "file_name":
                part.get_filename() or "",
            "content_type":
                part.get_content_type(),
            "size_bytes":
                len(payload),
        })

    headers = {
        "subject":
            normalize_text(
                message.get("subject")
            ),
        "from":
            normalize_text(
                message.get("from")
            ),
        "to":
            normalize_text(
                message.get("to")
            ),
        "cc":
            normalize_text(
                message.get("cc")
            ),
        "date":
            normalize_text(
                message.get("date")
            ),
        "message_id":
            normalize_text(
                message.get("message-id")
            ),
    }

    records = []

    for index, chunk in enumerate(
        _split_long_text(body),
        start=1,
    ):
        title = headers["subject"]

        if len(_split_long_text(body)) > 1:
            title = (
                f"{title} "
                f"({index})"
            )

        records.append(
            _record(
                record_type="email_body",
                title=title,
                content=chunk,
                sequence=index,
                metadata={
                    **headers,
                    "body_format": (
                        "plain"
                        if plain_body
                        else "html"
                    ),
                },
            )
        )

    metadata = {
        **headers,
        "attachments": attachments,
    }

    return records, metadata


def analyze_file(
    extension: str,
    content: bytes,
) -> tuple[list[dict], dict]:
    normalized_extension = (
        normalize_extension(extension)
    )

    if normalized_extension == "json":
        return _records_from_json(content), {}

    if normalized_extension == "xml":
        return _records_from_xml(content), {}

    if normalized_extension == "csv":
        return _records_from_csv(content), {}

    if normalized_extension == "xlsx":
        return _records_from_xlsx(content), {}

    if normalized_extension == "txt":
        return _records_from_txt(content), {}

    if normalized_extension == "pdf":
        return _records_from_pdf(content)

    if normalized_extension == "eml":
        return _records_from_eml(content)

    raise HTTPException(
        status_code=400,
        detail=(
            f".{normalized_extension}の"
            "解析処理は実装されていません。"
        ),
    )


def validate_supported_extension(
    extension: str,
) -> str:
    normalized_extension = (
        normalize_extension(extension)
    )

    if not normalized_extension:
        raise HTTPException(
            status_code=400,
            detail="拡張子を確認できません。",
        )

    document = (
        get_firestore_client()
        .collection(
            SUPPORTED_FILE_TYPE_COLLECTION
        )
        .document(normalized_extension)
        .get()
    )

    if not document.exists:
        raise HTTPException(
            status_code=400,
            detail=(
                f".{normalized_extension}は"
                "拡張子管理に登録されていません。"
            ),
        )

    data = document.to_dict() or {}

    if data.get("enabled", True) is False:
        raise HTTPException(
            status_code=400,
            detail=(
                f".{normalized_extension}は"
                "無効になっています。"
            ),
        )

    return normalized_extension


def get_source_file(
    data_source_id: str,
    source_id: str,
) -> dict:
    normalized_data_source_id = normalize_text(
        data_source_id
    )
    normalized_source_id = normalize_text(
        source_id
    )

    if not normalized_data_source_id:
        raise HTTPException(
            status_code=400,
            detail="データソースIDが指定されていません。",
        )

    if not normalized_source_id:
        raise HTTPException(
            status_code=400,
            detail="解析対象IDが指定されていません。",
        )

    reference = (
        get_firestore_client()
        .collection(DATA_SOURCE_COLLECTION)
        .document(normalized_data_source_id)
        .collection(DATA_IMPORT_COLLECTION)
        .document(normalized_source_id)
    )
    document = reference.get()

    if not document.exists:
        raise HTTPException(
            status_code=404,
            detail="解析対象ファイルが見つかりません。",
        )

    data = document.to_dict() or {}

    if data.get("deleted", False):
        raise HTTPException(
            status_code=400,
            detail="削除済みのファイルです。",
        )

    gcs_path = normalize_text(data.get("gcs_path"))
    extension = normalize_extension(data.get("extension"))

    if not extension:
        file_name = normalize_text(data.get("file_name"))
        if "." in file_name:
            extension = normalize_extension(
                PurePosixPath(file_name).suffix
            )

    if not gcs_path:
        raise HTTPException(
            status_code=400,
            detail="Cloud Storageの保存先が登録されていません。",
        )

    return {
        "source_type": "data_import",
        "source_id": normalized_source_id,
        "source_reference": reference,
        "data_source_id": normalized_data_source_id,
        "data_source_name": normalize_text(
            data.get("data_source_name")
        ),
        "tenant_id": normalize_text(data.get("tenant_id")),
        "file_name": (
            normalize_text(data.get("file_name"))
            or PurePosixPath(gcs_path).name
        ),
        "content_type": normalize_text(
            data.get("content_type")
        ),
        "extension": extension,
        "bucket_name": normalize_text(
            data.get("bucket_name")
        ),
        "gcs_path": gcs_path,
        "gcs_uri": normalize_text(data.get("gcs_uri")),
    }


def _download_source(
    source: dict,
) -> bytes:
    try:
        blob = get_storage_bucket().blob(
            source["gcs_path"]
        )

        if not blob.exists():
            raise HTTPException(
                status_code=404,
                detail=(
                    "Cloud Storageに"
                    "ファイルがありません。"
                ),
            )

        return blob.download_as_bytes()
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(
            status_code=500,
            detail=(
                "Cloud Storageから"
                "ファイルを取得できません。"
                f" {type(error).__name__}: {error}"
            ),
        )



def get_raw_document_reference(
    db,
    data_source_id: str,
    document_id: str,
):
    normalized_data_source_id = normalize_text(
        data_source_id
    )
    normalized_document_id = normalize_text(
        document_id
    )

    if not normalized_data_source_id:
        raise HTTPException(
            status_code=400,
            detail="データソースIDがありません。",
        )

    if not normalized_document_id:
        raise HTTPException(
            status_code=400,
            detail="元データIDがありません。",
        )

    return (
        db.collection(DATA_SOURCE_COLLECTION)
        .document(normalized_data_source_id)
        .collection(RAW_DOCUMENT_COLLECTION)
        .document(normalized_document_id)
    )


def _delete_existing_records(
    document_reference,
) -> None:
    db = get_firestore_client()

    while True:
        documents = list(
            document_reference
            .collection(
                RAW_RECORD_SUBCOLLECTION
            )
            .limit(BATCH_WRITE_LIMIT)
            .stream()
        )

        if not documents:
            return

        batch = db.batch()

        for document in documents:
            batch.delete(
                document.reference
            )

        batch.commit()


def _write_records(
    document_reference,
    document_id: str,
    records: list[dict],
    source: dict,
    user: dict,
) -> None:
    db = get_firestore_client()
    batch = db.batch()
    operation_count = 0

    for index, record in enumerate(
        records,
        start=1,
    ):
        record_id = (
            f"{index:08d}_"
            f"{uuid.uuid4().hex[:8]}"
        )

        record_data = {
            "record_id":
                record_id,
            "document_id":
                document_id,
            "source_type":
                source["source_type"],
            "source_id":
                source["source_id"],
            "data_source_id":
                source["data_source_id"],
            "tenant_id":
                source["tenant_id"],
            **record,
            "created_at":
                now_iso(),
            "created_by":
                user.get("email", ""),
        }

        reference = (
            document_reference
            .collection(
                RAW_RECORD_SUBCOLLECTION
            )
            .document(record_id)
        )

        batch.set(
            reference,
            record_data,
        )

        operation_count += 1

        if operation_count >= BATCH_WRITE_LIMIT:
            batch.commit()
            batch = db.batch()
            operation_count = 0

    if operation_count:
        batch.commit()


def process_source_file(
    *,
    data_source_id: str,
    source_id: str,
    user: dict,
    overwrite: bool = True,
) -> dict:
    source = get_source_file(
        data_source_id,
        source_id,
    )

    extension = validate_supported_extension(
        source["extension"]
    )

    content = _download_source(source)

    records, analysis_metadata = analyze_file(
        extension,
        content,
    )

    document_id = (
        f"{source['source_type']}_"
        f"{source['source_id']}"
    )

    db = get_firestore_client()

    document_reference = get_raw_document_reference(
        db,
        source["data_source_id"],
        document_id,
    )

    existing_document = (
        document_reference.get()
    )

    if existing_document.exists and not overwrite:
        raise HTTPException(
            status_code=409,
            detail=(
                "このファイルは既に"
                "Firestoreへ登録されています。"
            ),
        )

    if existing_document.exists:
        _delete_existing_records(
            document_reference
        )

    now = now_iso()

    document_data = {
        "document_id":
            document_id,
        "source_type":
            source["source_type"],
        "source_id":
            source["source_id"],
        "data_source_id":
            source["data_source_id"],
        "data_source_name":
            source["data_source_name"],
        "tenant_id":
            source["tenant_id"],
        "file_name":
            source["file_name"],
        "extension":
            extension,
        "content_type":
            source["content_type"],
        "bucket_name":
            source["bucket_name"],
        "gcs_path":
            source["gcs_path"],
        "gcs_uri":
            source["gcs_uri"],
        "size_bytes":
            len(content),
        "record_count":
            len(records),
        "analysis_metadata":
            _safe_value(
                analysis_metadata
            ),
        "status":
            "processed",
        "processed_at":
            now,
        "processed_by":
            user.get("email", ""),
        "updated_at":
            now,
        "updated_by":
            user.get("email", ""),
    }

    if not existing_document.exists:
        document_data.update({
            "created_at":
                now,
            "created_by":
                user.get("email", ""),
        })

    document_reference.set(
        document_data,
        merge=True,
    )

    try:
        _write_records(
            document_reference,
            document_id,
            records,
            source,
            user,
        )
    except Exception as error:
        document_reference.set({
            "status":
                "failed",
            "error_message":
                str(error),
            "updated_at":
                now_iso(),
        }, merge=True)

        raise

    source_reference = source["source_reference"]

    source_reference.set({
        "raw_document_id":
            document_id,
        "analysis_status":
            "completed",
        "analysis_record_count":
            len(records),
        "analyzed_at":
            now_iso(),
        "analysis_completed_at":
            now_iso(),
        "analyzed_by":
            user.get("email", ""),
    }, merge=True)

    return {
        "status":
            "completed",
        "document_id":
            document_id,
        "source_type":
            source["source_type"],
        "source_id":
            source["source_id"],
        "extension":
            extension,
        "record_count":
            len(records),
        "split_method":
            analysis_metadata.get(
                "split_method"
            ),
    }
