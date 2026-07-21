import io
import json

from fastapi import HTTPException
from openpyxl import load_workbook

from app.routers.data_raw_analysis_common import (
    create_record,
    safe_value,
)
from app.routers.data_import_common import (
    normalize_text,
)


def analyze_xlsx(
    content: bytes,
) -> tuple[list[dict], dict]:
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
        )
    except Exception as error:
        raise HTTPException(
            status_code=400,
            detail=(
                "XLSXファイルを解析できません。"
                f" {type(error).__name__}: {error}"
            ),
        )

    records = []
    sequence = 0
    sheet_names = []

    try:
        for worksheet in workbook.worksheets:
            sheet_names.append(
                worksheet.title
            )

            rows = worksheet.iter_rows(
                values_only=True
            )

            try:
                first_row = next(rows)
            except StopIteration:
                continue

            headers = [
                normalize_text(value)
                or f"column_{index}"
                for index, value in enumerate(
                    first_row,
                    start=1,
                )
            ]

            for row_number, values in enumerate(
                rows,
                start=2,
            ):
                if not any(
                    value is not None
                    and normalize_text(value)
                    for value in values
                ):
                    continue

                sequence += 1

                row_data = {
                    headers[index]:
                        safe_value(value)
                    for index, value in enumerate(
                        values
                    )
                    if index < len(headers)
                }

                records.append(
                    create_record(
                        record_type=
                            "xlsx_row",
                        title=(
                            f"{worksheet.title}"
                            f" / 行 {row_number}"
                        ),
                        sequence=
                            sequence,
                        structured_data=
                            row_data,
                        content=json.dumps(
                            row_data,
                            ensure_ascii=False,
                            separators=(",", ":"),
                        ),
                        metadata={
                            "sheet_name":
                                worksheet.title,
                            "row_number":
                                row_number,
                            "headers":
                                headers,
                        },
                    )
                )
    finally:
        workbook.close()

    return records, {
        "sheet_names": sheet_names,
    }
